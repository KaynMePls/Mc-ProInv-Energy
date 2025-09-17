from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file
from flask_login import login_required, current_user
from app.models import Producto, CategoriaComponente, Usuario, MovimientoStock
from app.forms import ProductoForm
from app import db
from werkzeug.utils import secure_filename
from io import BytesIO
from datetime import datetime, timedelta
import io
import os
import pandas as pd
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from collections import Counter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime



inventarios_bp = Blueprint('inventarios', __name__, url_prefix='/inventarios')

# Dashboard de inventarios: muestra solo productos publicados
@inventarios_bp.route('/dashboard')
@login_required
def dashboard():
    productos = Producto.query.filter_by(publicado=True).all()
    total_productos = len(productos)
    productos_stock_bajo = sum(
        1 for p in productos
        if p.stock is not None and p.stock_minimo is not None and p.stock <= p.stock_minimo
    )
    hace_7_dias = datetime.utcnow() - timedelta(days=7)
    productos_nuevos = sum(
        1 for p in productos
        if p.fecha_actualizacion and p.fecha_actualizacion >= hace_7_dias
    )
    valor_inventario = sum(
        float(p.precio or 0) * float(p.stock or 0)
        for p in productos
    )
    ultimos_productos = sorted(
        productos, key=lambda p: p.fecha_actualizacion or datetime.min, reverse=True
    )[:5]
    categorias_stock = []
    stock_por_categoria = []
    categorias_vistas = set()
    for p in productos:
        if p.categoria and p.categoria.nombre not in categorias_vistas:
            categorias_vistas.add(p.categoria.nombre)
            categorias_stock.append(p.categoria.nombre)
            stock_categoria = sum(
                prod.stock or 0 for prod in productos
                if prod.categoria and prod.categoria.nombre == p.categoria.nombre
            )
            stock_por_categoria.append(stock_categoria)
    return render_template(
        'empleado/inventarios_dashboard.html',
        productos=productos,
        total_productos=total_productos,
        productos_stock_bajo=productos_stock_bajo,
        productos_nuevos=productos_nuevos,
        valor_inventario=valor_inventario,
        ultimos_productos=ultimos_productos,
        categorias_stock=categorias_stock,
        stock_por_categoria=stock_por_categoria
    )

# Listado de productos
@inventarios_bp.route('/productos')
@login_required
def productos():
    buscar = request.args.get('buscar', '').strip()
    categoria_nombre = request.args.get('categoria', '')
    stock_filtro = request.args.get('stock', '')
    publicado_filtro = request.args.get('publicado', '')

    query = Producto.query
    if publicado_filtro == 'si':
        query = query.filter_by(publicado=True)
    elif publicado_filtro == 'no':
        query = query.filter_by(publicado=False)
    # Si no se especifica, muestra todos

    if buscar:
        query = query.join(CategoriaComponente, isouter=True).filter(
            (Producto.nombre.ilike(f'%{buscar}%')) |
            (CategoriaComponente.nombre.ilike(f'%{buscar}%'))
        )
    if categoria_nombre:
        query = query.join(CategoriaComponente).filter(CategoriaComponente.nombre == categoria_nombre)
    if stock_filtro == 'bajo':
        query = query.filter(Producto.stock < 5)
    elif stock_filtro == 'alto':
        query = query.filter(Producto.stock >= 5)

    productos = query.all()
    categorias = CategoriaComponente.query.all()
    return render_template(
        'empleado/inventarios_productos.html',
        productos=productos,
        categorias=categorias
    )

# Crear nuevo producto (por personal autorizado, no proveedores)
@inventarios_bp.route('/producto/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_producto():
    categorias = CategoriaComponente.query.all()
    proveedores = Usuario.query.filter_by(rol='proveedor').all()
    sku_sugerido = "SKU-" + str(Producto.query.count() + 1)

    if request.method == 'POST':
        nombre = request.form['nombre']
        descripcion = request.form['descripcion']
        categoria_id = request.form['categoria']
        stock = request.form['stock']
        precio = request.form['precio']
        proveedor_id = request.form['proveedor_id']
        ubicacion = request.form['ubicacion']
        estado = request.form['estado']
        sku = request.form['codigo']
        imagen = request.files['imagen']

        # Limpiar y convertir precio
        precio_limpio = float(precio.replace('.', '').replace(',', '.'))

        # Guardar imagen en la carpeta correcta y guardar solo el nombre
        imagen_url = None
        if imagen and imagen.filename != '':
            filename = secure_filename(imagen.filename)
            ruta_upload = os.path.join('app', 'static', 'img_productos', filename)
            os.makedirs(os.path.dirname(ruta_upload), exist_ok=True)
            imagen.save(ruta_upload)
            imagen_url = filename  # Solo el nombre

        producto = Producto(
            nombre=nombre,
            descripcion=descripcion,
            categoria_id=categoria_id,
            stock=stock,
            precio=precio_limpio,
            proveedor_id=proveedor_id,
            ubicacion=ubicacion,
            estado=estado,
            sku=sku,
            imagen_url=imagen_url,
            publicado=True
        )
        db.session.add(producto)
        db.session.commit()
        flash('Producto agregado exitosamente.', 'success')
        return redirect(url_for('inventarios.productos'))

    cancel_url = request.args.get('cancel_url') or url_for('inventarios.productos')
    return render_template(
        'empleado/inventarios_agregar.html',
        categorias=categorias,
        proveedores=proveedores,
        sku_sugerido=sku_sugerido,
        cancel_url=cancel_url
    )
# Editar producto existente
@inventarios_bp.route('/producto/editar/<int:producto_id>', methods=['GET', 'POST'])
@login_required
def editar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    categorias = CategoriaComponente.query.all()
    proveedores = Usuario.query.filter_by(rol='proveedor').all()

    cancel_url = request.args.get('cancel_url') or url_for('inventarios.productos')


    if request.method == 'POST':
        producto.nombre = request.form['nombre']
        producto.descripcion = request.form['descripcion']
        producto.categoria_id = request.form['categoria']
        producto.stock = request.form['stock']
        # Limpiar y convertir precio
        precio = request.form['precio']
        producto.precio = float(precio.replace('.', '').replace(',', '.'))
        producto.proveedor_id = request.form['proveedor']
        producto.ubicacion = request.form['ubicacion']
        producto.estado = request.form['estado']
        # producto.fecha_ingreso = request.form['fecha_ingreso'] # si aplica

        # Manejo de imagen
        imagen = request.files.get('imagen')
        if imagen and imagen.filename != '':
            filename = secure_filename(imagen.filename)
            ruta_upload = os.path.join('app', 'static', 'img_productos', filename)
            os.makedirs(os.path.dirname(ruta_upload), exist_ok=True)
            imagen.save(ruta_upload)
            producto.imagen_url = filename

        db.session.commit()
        flash('Producto actualizado correctamente.', 'success')
        return redirect(cancel_url)  

    return render_template(
        'empleado/inventarios_editar.html',
        producto=producto,
        categorias=categorias,
        proveedores=proveedores,
        cancel_url=cancel_url
    )
        

# Editar solo el stock de un producto
@inventarios_bp.route('/producto/editar_stock/<int:producto_id>', methods=['GET', 'POST'])
@login_required
def editar_stock(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    if request.method == 'POST':
        nuevo_stock = request.form.get('stock')
        if nuevo_stock is not None and nuevo_stock.isdigit():
            producto.stock = int(nuevo_stock)
            db.session.commit()
            flash('Stock actualizado correctamente.', 'success')
            return redirect(url_for('inventarios.productos'))
        else:
            flash('Stock inválido.', 'danger')
    return render_template('empleado/inventarios_editar_stock.html', producto=producto)

# Detalle de producto
@inventarios_bp.route('/producto/<int:id>')
@login_required
def detalle_producto(id):
    producto = Producto.query.get_or_404(id)
    movimientos = MovimientoStock.query.filter_by(producto_id=producto.id).order_by(MovimientoStock.fecha.desc()).all()
    return render_template(
        'empleado/inventarios_detalle.html',
        producto=producto,
        movimientos=movimientos
    )

# Reporte de inventario 
@inventarios_bp.route('/reporte')
@login_required
def reporte():
    productos = Producto.query.filter_by(publicado=True).all()
    categorias = CategoriaComponente.query.all()
    usuarios = Usuario.query.all()
    total_productos = len(productos)
    valor_inventario = sum(float(p.precio or 0) * float(p.stock or 0) for p in productos)
    total_categorias = len({p.categoria_id for p in productos if p.categoria_id})
    total_proveedores = len({p.proveedor_id for p in productos if p.proveedor_id})
    productos_stock_bajo = sum(1 for p in productos if p.stock is not None and p.stock_minimo is not None and p.stock <= p.stock_minimo)
    productos_agotados = sum(1 for p in productos if p.stock == 0)
    productos_sin_proveedor = sum(1 for p in productos if not p.proveedor_id)
    productos_sin_categoria = sum(1 for p in productos if not p.categoria_id)
    producto_mayor_stock = max(productos, key=lambda p: p.stock or 0, default=None)
    producto_menor_stock = min(productos, key=lambda p: p.stock if p.stock is not None else float('inf'), default=None)
    producto_mas_caro = max(productos, key=lambda p: p.precio or 0, default=None)
    producto_mas_barato = min(productos, key=lambda p: p.precio if p.precio is not None else float('inf'), default=None)
    categoria_counter = Counter([p.categoria_id for p in productos if p.categoria_id])
    categoria_top_id, categoria_top_count = (categoria_counter.most_common(1)[0] if categoria_counter else (None, 0))
    categoria_top = next((c for c in categorias if c.id == categoria_top_id), None)
    if categoria_top:
        categoria_top.cantidad = categoria_top_count

    proveedor_counter = Counter([p.proveedor_id for p in productos if p.proveedor_id])
    proveedor_top_id, proveedor_top_count = (proveedor_counter.most_common(1)[0] if proveedor_counter else (None, 0))
    proveedor_top = next((u for u in usuarios if u.id == proveedor_top_id), None)
    if proveedor_top:
        proveedor_top.cantidad = proveedor_top_count
    productos_criticos = [p for p in productos if (p.stock is not None and p.stock <= (p.stock_minimo or 5))]
    recomendaciones = []
    if productos_stock_bajo > 0:
        recomendaciones.append("Revisar y reabastecer los productos con stock bajo para evitar quiebres de inventario.")
    if productos_agotados > 0:
        recomendaciones.append("Solicitar reposición urgente de productos agotados.")
    if productos_sin_proveedor > 0:
        recomendaciones.append("Asignar proveedor a los productos que no lo tienen.")
    if productos_sin_categoria > 0:
        recomendaciones.append("Clasificar los productos sin categoría para mejorar la gestión.")
    if not recomendaciones:
        recomendaciones.append("El inventario se encuentra en condiciones óptimas.")
    fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")
    return render_template(
        'empleado/inventarios_reporte.html',
        fecha_reporte=fecha_reporte,
        total_productos=total_productos,
        valor_inventario=valor_inventario,
        total_categorias=total_categorias,
        total_proveedores=total_proveedores,
        productos_stock_bajo=productos_stock_bajo,
        productos_agotados=productos_agotados,
        productos_sin_proveedor=productos_sin_proveedor,
        productos_sin_categoria=productos_sin_categoria,
        producto_mayor_stock=producto_mayor_stock,
        producto_menor_stock=producto_menor_stock,
        producto_mas_caro=producto_mas_caro,
        producto_mas_barato=producto_mas_barato,
        categoria_top=categoria_top,
        proveedor_top=proveedor_top,
        productos_criticos=productos_criticos,
        recomendaciones=recomendaciones,
        usuarios=usuarios
    )

# Reporte manual 
@inventarios_bp.route('/reporte_manual')
@login_required
def reporte_manual():
    productos = Producto.query.filter_by(publicado=True).all()
    total_productos = len(productos)
    valor_inventario = sum(float(p.precio or 0) * float(p.stock or 0) for p in productos)
    total_categorias = len({p.categoria_id for p in productos if p.categoria_id})
    total_proveedores = len({p.proveedor_id for p in productos if p.proveedor_id})
    productos_stock_bajo = sum(1 for p in productos if p.stock is not None and p.stock_minimo is not None and p.stock <= p.stock_minimo)
    productos_agotados = sum(1 for p in productos if p.stock == 0)
    productos_criticos = [p for p in productos if (p.stock is not None and p.stock <= (p.stock_minimo or 5))]
    recomendaciones = []
    if productos_stock_bajo > 0:
        recomendaciones.append("Revisar y reabastecer los productos con stock bajo para evitar quiebres de inventario.")
    if productos_agotados > 0:
        recomendaciones.append("Solicitar reposición urgente de productos agotados.")
    if not recomendaciones:
        recomendaciones.append("El inventario se encuentra en condiciones óptimas.")
    responsable = current_user.nombre if hasattr(current_user, 'nombre') else current_user.username
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    return render_template(
        'empleado/inventarios.reporte_manual.html',
        total_productos=total_productos,
        valor_inventario=valor_inventario,
        total_categorias=total_categorias,
        total_proveedores=total_proveedores,
        productos_stock_bajo=productos_stock_bajo,
        productos_agotados=productos_agotados,
        recomendaciones=recomendaciones,
        productos_criticos=productos_criticos,
        responsable=responsable,
        fecha_actual=fecha_actual
    )

@inventarios_bp.route('/exportar_excel')
@login_required
def exportar_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter

    buscar = request.args.get('buscar', '').strip()
    categoria_nombre = request.args.get('categoria', '')
    stock_filtro = request.args.get('stock', '')

    query = Producto.query.filter_by(publicado=True)
    if buscar:
        query = query.join(CategoriaComponente, isouter=True).filter(
            (Producto.nombre.ilike(f'%{buscar}%')) |
            (CategoriaComponente.nombre.ilike(f'%{buscar}%'))
        )
    if categoria_nombre:
        query = query.join(CategoriaComponente).filter(CategoriaComponente.nombre == categoria_nombre)
    if stock_filtro == 'bajo':
        query = query.filter(Producto.stock < 5)
    elif stock_filtro == 'alto':
        query = query.filter(Producto.stock >= 5)

    productos = query.all()
    headers = [
        'Nombre', 'Categoría', 'Stock', 'Precio', 'Proveedor',
        'Fecha de Actualización', 'SKU', 'Ubicación', 'Estado'
    ]
    data = []
    for p in productos:
        data.append([
            p.nombre,
            p.categoria.nombre if p.categoria else '',
            p.stock,
            float(p.precio or 0),
            p.proveedor.nombre if p.proveedor else '',
            p.fecha_actualizacion.strftime('%Y-%m-%d %H:%M') if p.fecha_actualizacion else '',
            p.sku,
            p.ubicacion,
            p.estado,
        ])

    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1976D2")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="38404a")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Escribir datos
    for row_num, row in enumerate(data, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(border_style="thin", color="38404a")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            # Formato moneda para precio
            if headers[col_num-1] == "Precio":
                cell.number_format = '"$"#,##0.00'

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Filtros en la primera fila
    ws.auto_filter.ref = ws.dimensions

    # Congelar encabezados
    ws.freeze_panes = "A2"

    # Guardar en memoria y enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="inventario_profesional.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@inventarios_bp.route('/exportar_pdf')
@login_required
def exportar_pdf():
 
    buscar = request.args.get('buscar', '').strip()
    categoria_nombre = request.args.get('categoria', '')
    stock_filtro = request.args.get('stock', '')

    query = Producto.query.filter_by(publicado=True)
    if buscar:
        query = query.join(CategoriaComponente, isouter=True).filter(
            (Producto.nombre.ilike(f'%{buscar}%')) |
            (CategoriaComponente.nombre.ilike(f'%{buscar}%'))
        )
    if categoria_nombre:
        query = query.join(CategoriaComponente).filter(CategoriaComponente.nombre == categoria_nombre)
    if stock_filtro == 'bajo':
        query = query.filter(Producto.stock < 5)
    elif stock_filtro == 'alto':
        query = query.filter(Producto.stock >= 5)

    productos = query.all()
    headers = ["Nombre", "Categoría", "Stock", "Precio", "Proveedor", "Fecha de Actualización", "SKU", "Ubicación", "Estado", "Descripción"]

    # Usar Paragraph para descripción y otros textos largos
    styles = getSampleStyleSheet()
    desc_style = ParagraphStyle('desc_style', parent=styles['Normal'], fontSize=8, leading=10, alignment=0)
    cell_style = ParagraphStyle('cell_style', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)

    data = [headers]
    for p in productos:
        data.append([
            Paragraph(p.nombre or '', cell_style),
            Paragraph(p.categoria.nombre if p.categoria else '', cell_style),
            str(p.stock),
            f"${float(p.precio or 0):,.2f}",
            Paragraph(p.proveedor.nombre if p.proveedor else '', cell_style),
            p.fecha_actualizacion.strftime('%Y-%m-%d %H:%M') if p.fecha_actualizacion else '',
            p.sku or '',
            Paragraph(p.ubicacion or '', cell_style),
            p.estado or '',
            Paragraph(p.descripcion or '', desc_style)
        ])

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        leftMargin=20,
        rightMargin=20,
        topMargin=30,
        bottomMargin=20
    )
    elements = []
    title = Paragraph("<b>Reporte de Inventario</b>", styles['Title'])
    fecha = Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    elements.append(title)
    elements.append(fecha)
    elements.append(Spacer(1, 12))

    # Solo fija el ancho de la columna de descripción, deja las demás automáticas
    col_widths = [60, 60, 30, 50, 60, 70, 50, 60, 40, 180]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-2,-1), 'CENTER'),
        ('ALIGN', (-1,1), (-1,-1), 'LEFT'),  # Descripción alineada a la izquierda
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#38404a')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey])
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return send_file(output, download_name="inventario_profesional.pdf", as_attachment=True, mimetype='application/pdf')

# Publicar producto (solo para personal autorizado)
@inventarios_bp.route('/publicar/<int:producto_id>', methods=['POST'])
@login_required
def publicar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    producto.publicado = True
    db.session.commit()
    flash('Producto publicado exitosamente.', 'success')
    return redirect(request.referrer or url_for('inventarios.productos'))

# eliminar producto (solo para personal autorizado)
@inventarios_bp.route('/eliminar/<int:producto_id>', methods=['POST'])
@login_required
def eliminar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    db.session.delete(producto)
    db.session.commit()
    flash('Producto eliminado correctamente.', 'success')
    return redirect(url_for('inventarios.productos'))

@inventarios_bp.route('/exportar_pdf_reporte_general')
@login_required
def exportar_pdf_reporte_general():
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    productos = Producto.query.filter_by(publicado=True).all()
    categorias = CategoriaComponente.query.all()
    usuarios = Usuario.query.all()
    total_productos = len(productos)
    valor_inventario = sum(float(p.precio or 0) * float(p.stock or 0) for p in productos)
    total_categorias = len({p.categoria_id for p in productos if p.categoria_id})
    total_proveedores = len({p.proveedor_id for p in productos if p.proveedor_id})
    productos_stock_bajo = sum(1 for p in productos if p.stock is not None and p.stock_minimo is not None and p.stock <= p.stock_minimo)
    productos_agotados = sum(1 for p in productos if p.stock == 0)
    productos_sin_proveedor = sum(1 for p in productos if not p.proveedor_id)
    productos_sin_categoria = sum(1 for p in productos if not p.categoria_id)
    producto_mayor_stock = max(productos, key=lambda p: p.stock or 0, default=None)
    producto_menor_stock = min(productos, key=lambda p: p.stock if p.stock is not None else float('inf'), default=None)
    producto_mas_caro = max(productos, key=lambda p: p.precio or 0, default=None)
    producto_mas_barato = min(productos, key=lambda p: p.precio if p.precio is not None else float('inf'), default=None)
    from collections import Counter
    categoria_counter = Counter([p.categoria_id for p in productos if p.categoria_id])
    categoria_top_id, categoria_top_count = (categoria_counter.most_common(1)[0] if categoria_counter else (None, 0))
    categoria_top = next((c for c in categorias if c.id == categoria_top_id), None)
    if categoria_top:
        categoria_top.cantidad = categoria_top_count

    proveedor_counter = Counter([p.proveedor_id for p in productos if p.proveedor_id])
    proveedor_top_id, proveedor_top_count = (proveedor_counter.most_common(1)[0] if proveedor_counter else (None, 0))
    proveedor_top = next((u for u in usuarios if u.id == proveedor_top_id), None)
    if proveedor_top:
        proveedor_top.cantidad = proveedor_top_count
    productos_criticos = [p for p in productos if (p.stock is not None and p.stock <= (p.stock_minimo or 5))]
    recomendaciones = []
    if productos_stock_bajo > 0:
        recomendaciones.append("Revisar y reabastecer los productos con stock bajo para evitar quiebres de inventario.")
    if productos_agotados > 0:
        recomendaciones.append("Solicitar reposición urgente de productos agotados.")
    if productos_sin_proveedor > 0:
        recomendaciones.append("Asignar proveedor a los productos que no lo tienen.")
    if productos_sin_categoria > 0:
        recomendaciones.append("Clasificar los productos sin categoría para mejorar la gestión.")
    if not recomendaciones:
        recomendaciones.append("El inventario se encuentra en condiciones óptimas.")
    fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Estilos
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('cell_style', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
    desc_style = ParagraphStyle('desc_style', parent=styles['Normal'], fontSize=8, leading=10, alignment=0)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        leftMargin=20,
        rightMargin=20,
        topMargin=30,
        bottomMargin=20
    )
    elements = []

    # Título y fecha
    title = Paragraph("<b>Reporte General de Inventario</b>", styles['Title'])
    fecha = Paragraph(f"Generado: {fecha_reporte}", styles['Normal'])
    elements.append(title)
    elements.append(fecha)
    elements.append(Spacer(1, 12))

    # KPIs y resumen
    resumen = Paragraph(
        f"<b>Total productos:</b> {total_productos} &nbsp;&nbsp; "
        f"<b>Valor inventario:</b> ${valor_inventario:,.0f} &nbsp;&nbsp; "
        f"<b>Categorías:</b> {total_categorias} &nbsp;&nbsp; "
        f"<b>Proveedores:</b> {total_proveedores} &nbsp;&nbsp; "
        f"<b>Stock bajo:</b> {productos_stock_bajo} &nbsp;&nbsp; "
        f"<b>Agotados:</b> {productos_agotados} &nbsp;&nbsp; "
        f"<b>Sin proveedor:</b> {productos_sin_proveedor} &nbsp;&nbsp; "
        f"<b>Sin categoría:</b> {productos_sin_categoria}",
        styles['Normal']
    )
    elements.append(resumen)
    elements.append(Spacer(1, 8))

    # Productos destacados
    destacados = []
    if producto_mayor_stock:
        destacados.append(f"Producto con mayor stock: {producto_mayor_stock.nombre} ({producto_mayor_stock.stock})")
    if producto_menor_stock:
        destacados.append(f"Producto con menor stock: {producto_menor_stock.nombre} ({producto_menor_stock.stock})")
    if producto_mas_caro:
        destacados.append(f"Producto más caro: {producto_mas_caro.nombre} (${producto_mas_caro.precio:,.2f})")
    if producto_mas_barato:
        destacados.append(f"Producto más barato: {producto_mas_barato.nombre} (${producto_mas_barato.precio:,.2f})")
    if categoria_top:
        destacados.append(f"Categoría con más productos: {categoria_top.nombre} ({categoria_top.cantidad})")
    if proveedor_top:
        destacados.append(f"Proveedor con más productos: {proveedor_top.nombre} ({proveedor_top.cantidad})")
    if destacados:
        elements.append(Paragraph("<b>Indicadores destacados</b>", styles['Heading3']))
        for d in destacados:
            elements.append(Paragraph(f"- {d}", styles['Normal']))
        elements.append(Spacer(1, 8))

    # Tabla de productos críticos
    elements.append(Paragraph("<b>Productos críticos</b>", styles['Heading3']))
    headers = ["Nombre", "Categoría", "Proveedor", "Stock", "Estado"]
    data = [headers]
    for p in productos_criticos:
        data.append([
            Paragraph(p.nombre or '', cell_style),
            Paragraph(p.categoria.nombre if p.categoria else 'Sin categoría', cell_style),
            Paragraph(p.proveedor.nombre if p.proveedor else 'Sin proveedor', cell_style),
            str(p.stock),
            Paragraph(p.estado.capitalize() if p.estado else '', cell_style)
        ])
    col_widths = [100, 80, 80, 40, 60]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#38404a')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey])
    ]))
    elements.append(table)
    elements.append(Spacer(1, 18))

    # Recomendaciones
    elements.append(Paragraph("<b>Recomendaciones</b>", styles['Heading3']))
    for rec in recomendaciones:
        elements.append(Paragraph(f"- {rec}", styles['Normal']))
    elements.append(Spacer(1, 18))

    # Cierre profesional
    cierre = Paragraph(
        "Este reporte ha sido generado automáticamente por el sistema de gestión de inventarios MC Proinv.<br/>"
        "Para cualquier consulta adicional, contacte al área de operaciones o tecnología.",
        styles['Normal']
    )
    elements.append(cierre)

    doc.build(elements)
    output.seek(0)
    return send_file(output, download_name="reporte_general_inventario.pdf", as_attachment=True, mimetype='application/pdf')

@inventarios_bp.route('/exportar_excel_reporte_general')
@login_required
def exportar_excel_reporte_general():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    productos = Producto.query.filter_by(publicado=True).all()
    productos_criticos = [p for p in productos if (p.stock is not None and p.stock <= (p.stock_minimo or 5))]
    recomendaciones = []
    productos_stock_bajo = sum(1 for p in productos if p.stock is not None and p.stock_minimo is not None and p.stock <= p.stock_minimo)
    productos_agotados = sum(1 for p in productos if p.stock == 0)
    if productos_stock_bajo > 0:
        recomendaciones.append("Revisar y reabastecer los productos con stock bajo para evitar quiebres de inventario.")
    if productos_agotados > 0:
        recomendaciones.append("Solicitar reposición urgente de productos agotados.")
    if not recomendaciones:
        recomendaciones.append("El inventario se encuentra en condiciones óptimas.")

    headers = ["Nombre", "Categoría", "Proveedor", "Stock", "Estado"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos Críticos"

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1976D2")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="38404a")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Escribir datos
    for row_num, p in enumerate(productos_criticos, 2):
        ws.cell(row=row_num, column=1, value=p.nombre)
        ws.cell(row=row_num, column=2, value=p.categoria.nombre if p.categoria else 'Sin categoría')
        ws.cell(row=row_num, column=3, value=p.proveedor.nombre if p.proveedor else 'Sin proveedor')
        ws.cell(row=row_num, column=4, value=p.stock)
        ws.cell(row=row_num, column=5, value=p.estado.capitalize() if p.estado else '')

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Filtros en la primera fila
    ws.auto_filter.ref = ws.dimensions

    # Congelar encabezados
    ws.freeze_panes = "A2"

    # Hoja de recomendaciones
    ws2 = wb.create_sheet(title="Recomendaciones")
    ws2['A1'] = "Recomendaciones"
    ws2['A1'].font = Font(bold=True, color="1976D2")
    for idx, rec in enumerate(recomendaciones, 2):
        ws2[f'A{idx}'] = rec

    # Guardar en memoria y enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="reporte_general_inventario.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@inventarios_bp.route('/exportar_excel_reporte_manual')
@login_required
def exportar_excel_reporte_manual():
    # Obtén los datos igual que en la vista reporte_manual
    productos = Producto.query.filter_by(publicado=True).all()
    total_productos = len(productos)
    valor_inventario = sum(float(p.precio or 0) * float(p.stock or 0) for p in productos)
    total_categorias = len({p.categoria_id for p in productos if p.categoria_id})
    total_proveedores = len({p.proveedor_id for p in productos if p.proveedor_id})
    productos_stock_bajo = sum(1 for p in productos if p.stock is not None and p.stock_minimo is not None and p.stock <= p.stock_minimo)
    productos_agotados = sum(1 for p in productos if p.stock == 0)
    productos_criticos = [p for p in productos if (p.stock is not None and p.stock <= (p.stock_minimo or 5))]
    recomendaciones = []
    if productos_stock_bajo > 0:
        recomendaciones.append("Revisar y reabastecer los productos con stock bajo para evitar quiebres de inventario.")
    if productos_agotados > 0:
        recomendaciones.append("Solicitar reposición urgente de productos agotados.")
    if not recomendaciones:
        recomendaciones.append("El inventario se encuentra en condiciones óptimas.")
    responsable = current_user.nombre if hasattr(current_user, 'nombre') else current_user.username
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")

    headers = ["Nombre", "Categoría", "Proveedor", "Stock", "Estado"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos Críticos"

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1976D2")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="38404a")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Escribir datos
    for row_num, p in enumerate(productos_criticos, 2):
        ws.cell(row=row_num, column=1, value=p.nombre)
        ws.cell(row=row_num, column=2, value=p.categoria.nombre if p.categoria else 'Sin categoría')
        ws.cell(row=row_num, column=3, value=p.proveedor.nombre if p.proveedor else 'Sin proveedor')
        ws.cell(row=row_num, column=4, value=p.stock)
        ws.cell(row=row_num, column=5, value=p.estado.capitalize() if p.estado else '')

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Filtros en la primera fila
    ws.auto_filter.ref = ws.dimensions

    # Congelar encabezados
    ws.freeze_panes = "A2"

    # Hoja de resumen y recomendaciones
    ws2 = wb.create_sheet(title="Resumen y Recomendaciones")
    ws2['A1'] = "Responsable"
    ws2['B1'] = responsable
    ws2['A2'] = "Fecha"
    ws2['B2'] = fecha_actual
    ws2['A4'] = "Total productos"
    ws2['B4'] = total_productos
    ws2['A5'] = "Valor inventario"
    ws2['B5'] = valor_inventario
    ws2['A6'] = "Categorías"
    ws2['B6'] = total_categorias
    ws2['A7'] = "Proveedores"
    ws2['B7'] = total_proveedores
    ws2['A8'] = "Stock bajo"
    ws2['B8'] = productos_stock_bajo
    ws2['A9'] = "Agotados"
    ws2['B9'] = productos_agotados
    ws2['A11'] = "Recomendaciones"
    for idx, rec in enumerate(recomendaciones, 12):
        ws2[f'A{idx}'] = rec

    # Guardar en memoria y enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="reporte_manual_inventario.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@inventarios_bp.route('/reporte/manual/pdf', methods=['POST'])
@login_required
def exportar_pdf_reporte_manual():
    responsable = request.form.get('responsable', 'No especificado')
    fecha = request.form.get('fecha', '')
    resumen = {
        'total_productos': request.form.get('total_productos', ''),
        'valor_inventario': request.form.get('valor_inventario', ''),
        'total_categorias': request.form.get('total_categorias', ''),
        'total_proveedores': request.form.get('total_proveedores', ''),
        'productos_stock_bajo': request.form.get('productos_stock_bajo', ''),
        'productos_agotados': request.form.get('productos_agotados', ''),
    }
    recomendaciones = request.form.getlist('recomendaciones')
    if not recomendaciones or isinstance(recomendaciones, str):
        recomendaciones = request.form.get('recomendaciones', '').splitlines()
    from app.models import Producto
    productos_criticos = Producto.query.filter(Producto.stock < 5).all()
    reporte_manual = request.form.get('reporte_manual', '')

    buffer = BytesIO()
    width, height = letter

    # Usar platypus para mejor control de tablas y textos
    from reportlab.platypus import SimpleDocTemplate, Spacer
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    normal = styles['Normal']
    bold = styles['Heading2']
    bold.fontSize = 14

    # Título
    elements.append(Paragraph("<b>Reporte Manual de Inventario</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    # Responsable y fecha
    elements.append(Paragraph(f"<b>Responsable:</b> {responsable} &nbsp;&nbsp;&nbsp;&nbsp; <b>Fecha:</b> {fecha}", normal))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<hr width='100%'/>", normal))

    # Resumen Automático
    elements.append(Paragraph("<b>Resumen Automático:</b>", bold))
    resumen_html = (
        f"<b>Total productos:</b> {resumen['total_productos']}<br/>"
        f"<b>Valor inventario:</b> {resumen['valor_inventario']}<br/>"
        f"<b>Categorías:</b> {resumen['total_categorias']}<br/>"
        f"<b>Proveedores:</b> {resumen['total_proveedores']}<br/>"
        f"<b>Productos con stock bajo:</b> {resumen['productos_stock_bajo']}<br/>"
        f"<b>Productos agotados:</b> {resumen['productos_agotados']}"
    )
    elements.append(Paragraph(resumen_html, normal))
    elements.append(Spacer(1, 10))

    # Recomendaciones automáticas
    if recomendaciones and any(r.strip() for r in recomendaciones):
        elements.append(Paragraph("<b>Recomendaciones automáticas:</b>", bold))
        for rec in recomendaciones:
            if rec.strip():
                elements.append(Paragraph(f"- {rec}", normal))
        elements.append(Spacer(1, 10))

    # Productos críticos (tabla)
    elements.append(Paragraph("<b>Productos críticos:</b>", bold))
    table_data = [[
        "Nombre", "Categoría", "Proveedor", "Stock", "Estado", "Precio"
    ]]
    cell_style = ParagraphStyle('cell_style', parent=styles['Normal'], fontSize=9, leading=11)
    for prod in productos_criticos:
        table_data.append([
            Paragraph(prod.nombre, cell_style),
            Paragraph(prod.categoria.nombre if prod.categoria else "Sin categoría", cell_style),
            Paragraph(str(prod.proveedor_id), cell_style),
            Paragraph(str(prod.stock), cell_style),
            Paragraph(prod.estado, cell_style),
            Paragraph(f"${prod.precio:,.0f}".replace(",", "."), cell_style),
        ])
    col_widths = [90, 80, 60, 40, 50, 60]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1976d2")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 18))

    # Redacción del reporte manual
    elements.append(Paragraph("<b>Redacción del reporte manual:</b>", bold))
    for line in reporte_manual.splitlines():
        elements.append(Paragraph(line, normal))
    elements.append(Spacer(1, 12))

    # Pie de página
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<font size=9><i>Generado por MC-PROINV ENERGY</i></font>", normal))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="reporte_manual.pdf", mimetype='application/pdf')